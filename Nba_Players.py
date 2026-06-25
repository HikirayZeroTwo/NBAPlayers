import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl import load_workbook
import os

EXCEL_FILE = "NBA_Players.xlsx"
HEADERS = [
    "Имя", "Фамилия", "Возраст", "Рост (см)",
    "Текущая команда", "Количество матчей",
    "Игровые минуты (ср.)", "Очков за игру (ср.)",
    "Эффективность бросков (%)", "Трёхочковые (%)", "Штрафные (%)"
]

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Игроки NBA"
        ws.append(HEADERS)
        # Стиль заголовка
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1A478A")
        for col, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = 22
        wb.save(EXCEL_FILE)

def save_to_excel(data):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append(data)
    from openpyxl.styles import Alignment
    row = ws.max_row
    for cell in ws[row]:
        cell.alignment = Alignment(horizontal="center")
    wb.save(EXCEL_FILE)

def load_from_excel():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows.append(row)
    return rows

class NBAApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Игроки NBA")
        self.geometry("960x680")
        self.resizable(True, True)
        self.configure(bg="#F0F4FA")
        init_excel()
        self._build_ui()
        self._load_table()

    def _build_ui(self):
        # ── Заголовок ──────────────────────────────────────────────
        hdr = tk.Frame(self, bg="#1A478A", height=60)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🏀  Игроки NBA", font=("Segoe UI", 18, "bold"),
                 bg="#1A478A", fg="white").pack(pady=12)

        # ── Основная область ───────────────────────────────────────
        main = tk.Frame(self, bg="#F0F4FA")
        main.pack(fill="both", expand=True, padx=16, pady=12)

        # Левая панель — форма
        left = tk.LabelFrame(main, text="  Данные игрока  ",
                             font=("Segoe UI", 10, "bold"),
                             bg="#F0F5FA", fg="#1A478A",
                             relief="groove", bd=2)
        left.pack(side="left", fill="y", padx=(0, 12))

        fields_def = [
            ("Имя",                       "entry",  None),
            ("Фамилия",                   "entry",  None),
            ("Возраст",                   "entry",  None),
            ("Рост (см)",                 "entry",  None),
            ("Текущая команда",           "entry",  None),
            ("Количество матчей",         "entry",  None),
            ("Игровые минуты (среднее)",  "entry",  None),
            ("Очков за игру (среднее)",   "entry",  None),
            ("Эффективность бросков (%)", "entry",  None),
            ("Трёхочковые (%)",           "entry",  None),
            ("Штрафные (%)",              "entry",  None),
        ]

        self.entries = {}
        for i, (label, wtype, _) in enumerate(fields_def):
            tk.Label(left, text=label, font=("Segoe UI", 9),
                     bg="#000000", anchor="w").grid(
                row=i, column=0, sticky="w", padx=12, pady=4)
            e = tk.Entry(left, font=("Segoe UI", 10), width=22,
                         relief="solid", bd=1)
            e.grid(row=i, column=1, padx=(4, 12), pady=4)
            self.entries[label] = e

        # Кнопки
        btn_frame = tk.Frame(left, bg="#F0F4FA")
        btn_frame.grid(row=len(fields_def), column=0, columnspan=2, pady=12)

        tk.Button(btn_frame, text="💾  Добавить игрока",
                  font=("Segoe UI", 10, "bold"),
                  bg="#1A478A", fg="white", activebackground="#2460B0",
                  relief="flat", padx=16, pady=6,
                  command=self._add_player).pack(fill="x", pady=3)

        tk.Button(btn_frame, text="🗑  Очистить форму",
                  font=("Segoe UI", 9),
                  bg="#E0E8F4", fg="#1A478A", activebackground="#C8D8EC",
                  relief="flat", padx=16, pady=5,
                  command=self._clear_form).pack(fill="x", pady=3)

        tk.Button(btn_frame, text="❌  Удалить выбранного",
                  font=("Segoe UI", 9),
                  bg="#FAE0E0", fg="#A02020", activebackground="#F0C0C0",
                  relief="flat", padx=16, pady=5,
                  command=self._delete_player).pack(fill="x", pady=3)

        # Правая панель — таблица
        right = tk.LabelFrame(main, text="  Список игроков  ",
                              font=("Segoe UI", 10, "bold"),
                              bg="#F0F4FA", fg="#1A478A",
                              relief="groove", bd=2)
        right.pack(side="left", fill="both", expand=True)

        # Поиск
        search_frame = tk.Frame(right, bg="#F0F4FA")
        search_frame.pack(fill="x", padx=8, pady=(6, 2))
        tk.Label(search_frame, text="🔍 Поиск:", font=("Segoe UI", 9),
                 bg="#F0F4FA").pack(side="left")
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self._on_search)
        tk.Entry(search_frame, textvariable=self.search_var,
                 font=("Segoe UI", 9), width=20,
                 relief="solid", bd=1).pack(side="left", padx=6)

        # Таблица
        cols = ("№", "Имя", "Фамилия", "Команда", "Возраст",
                "Рост", "Матчи", "Мин.", "Очки", "FG%", "3P%", "FT%")
        self.tree = ttk.Treeview(right, columns=cols, show="headings",
                                 selectmode="browse")
        col_widths = [32, 80, 90, 110, 55, 55, 55, 55, 55, 55, 55, 55]
        for col, w in zip(cols, col_widths):
            self.tree.heading(col, text=col,
                              command=lambda c=col: self._sort(c))
            self.tree.column(col, width=w, anchor="center", minwidth=30)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", font=("Segoe UI", 9),
                        rowheight=26, background="#FFFFFF",
                        fieldbackground="#FFFFFF")
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"),
                        background="#1A478A", foreground="white")
        style.map("Treeview", background=[("selected", "#2460B0")])
        self.tree.tag_configure("odd", background="#F5F8FF")
        self.tree.tag_configure("even", background="#FFFFFF")

        vsb = ttk.Scrollbar(right, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(right, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=4)
        vsb.pack(side="right", fill="y", pady=4)
        hsb.pack(side="bottom", fill="x", padx=8)

        self.tree.bind("<<TreeviewSelect>>", self._on_select)

        # Статус бар
        self.status_var = tk.StringVar(value="Готово")
        tk.Label(self, textvariable=self.status_var,
                 font=("Segoe UI", 8), bg="#D8E4F0", fg="#1A478A",
                 anchor="w", padx=10).pack(fill="x", side="bottom")

    # ── Логика ────────────────────────────────────────────────────

    def _get_form_data(self):
        labels = list(self.entries.keys())
        data = []
        for lbl in labels:
            val = self.entries[lbl].get().strip()
            # Числовые поля — проверяем
            if lbl in ("Возраст", "Рост (см)", "Количество матчей"):
                if val and not val.isdigit():
                    messagebox.showwarning("Ошибка", f"Поле «{lbl}» должно быть целым числом.")
                    return None
                data.append(int(val) if val else "")
            elif lbl in ("Игровые минуты (среднее)", "Очков за игру (среднее)",
                         "Эффективность бросков (%)", "Трёхочковые (%)", "Штрафные (%)"):
                if val:
                    try:
                        data.append(float(val.replace(",", ".")))
                    except ValueError:
                        messagebox.showwarning("Ошибка", f"Поле «{lbl}» должно быть числом.")
                        return None
                else:
                    data.append("")
            else:
                data.append(val)
        return data

    def _add_player(self):
        data = self._get_form_data()
        if data is None:
            return
        if not data[0] and not data[1]:
            messagebox.showwarning("Ошибка", "Заполните хотя бы Имя или Фамилию.")
            return
        save_to_excel(data)
        self._load_table()
        self._clear_form()
        self.status_var.set(f"✅ Игрок {data[0]} {data[1]} добавлен в Excel.")

    def _clear_form(self):
        for e in self.entries.values():
            e.delete(0, tk.END)
        self.entries["Имя"].focus()

    def _delete_player(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Удаление", "Выберите игрока в таблице.")
            return
        item = self.tree.item(sel[0])
        row_num = item["values"][0]
        if not messagebox.askyesno("Удалить?",
                f"Удалить игрока {item['values'][1]} {item['values'][2]}?"):
            return
        # Удаляем из Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.delete_rows(int(row_num) + 1)  # +1 т.к. строка 1 — заголовок
        wb.save(EXCEL_FILE)
        self._load_table()
        self.status_var.set("🗑 Игрок удалён.")

    def _load_table(self, rows=None):
        self.tree.delete(*self.tree.get_children())
        if rows is None:
            rows = load_from_excel()
        for i, row in enumerate(rows, 1):
            tag = "odd" if i % 2 else "even"
            self.tree.insert("", "end",
                values=(i, *row), tags=(tag,))
        self.status_var.set(f"Загружено игроков: {len(rows)}")

    def _on_select(self, event):
        sel = self.tree.selection()
        if not sel:
            return
        vals = self.tree.item(sel[0])["values"]
        # vals: (№, Имя, Фамилия, Команда, Возраст, Рост, Матчи, Мин, Очки, FG%, 3P%, FT%)
        # Маппинг на поля формы
        mapping = [
            ("Имя", vals[1]),
            ("Фамилия", vals[2]),
            ("Возраст", vals[4]),
            ("Рост (см)", vals[5]),
            ("Текущая команда", vals[3]),
            ("Количество матчей", vals[6]),
            ("Игровые минуты (среднее)", vals[7]),
            ("Очков за игру (среднее)", vals[8]),
            ("Эффективность бросков (%)", vals[9]),
            ("Трёхочковые (%)", vals[10]),
            ("Штрафные (%)", vals[11]),
        ]
        for lbl, val in mapping:
            e = self.entries[lbl]
            e.delete(0, tk.END)
            if val is not None:
                e.insert(0, str(val))

    def _on_search(self, *args):
        query = self.search_var.get().lower()
        all_rows = load_from_excel()
        if not query:
            self._load_table(all_rows)
            return
        filtered = [r for r in all_rows
                    if any(query in str(v).lower() for v in r if v)]
        self._load_table(filtered)

    def _sort(self, col):
        cols = ("№", "Имя", "Фамилия", "Команда", "Возраст",
                "Рост", "Матчи", "Мин.", "Очки", "FG%", "3P%", "FT%")
        idx = cols.index(col)
        data = [(self.tree.set(c, col), c) for c in self.tree.get_children("")]
        try:
            data.sort(key=lambda x: float(x[0]) if x[0] else 0)
        except ValueError:
            data.sort(key=lambda x: x[0].lower())
        for i, (_, item) in enumerate(data):
            self.tree.move(item, "", i)
            tag = "odd" if i % 2 else "even"
            self.tree.item(item, tags=(tag,))

if __name__ == "__main__":
    app = NBAApp()
    app.mainloop()